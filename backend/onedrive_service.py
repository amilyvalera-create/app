"""Microsoft OneDrive for Business / Microsoft Graph data provider.

READ-ONLY. Reads the approved workbook worksheet "Precios Actual" via Microsoft
Graph (app-only client-credentials flow) and maps values strictly by the
confirmed column letters in schema.py. Until the corporate Microsoft credentials
are provided, it falls back to the schema-accurate MOCK dataset.

This layer is fully isolated: swapping to live data does NOT change auth, roles,
search, favorites or the UI.

REQUIRED SETUP (single pending item) — provide as secure env secrets:
  MS_TENANT_ID        Directory (tenant) ID (Entra / Azure app registration)
  MS_CLIENT_ID        Application (client) ID
  MS_CLIENT_SECRET    Client secret VALUE (not the secret id)
  ONEDRIVE_SHARE_URL  Full https://1drv.ms/... sharing link (already defaulted)
Required Graph application permission: Files.Read.All (admin-consented). Share
resolution may additionally require Files.ReadWrite.All per Graph docs.
"""

import os
import io
import base64
import logging

import httpx
from openpyxl import load_workbook

from schema import PRICE_COLUMNS, IDENTITY_COLUMNS, col_letter_to_index

logger = logging.getLogger("venege.onedrive")

WORKSHEET = "Precios Actual"
GRAPH = "https://graph.microsoft.com/v1.0"
DEFAULT_SHARE_URL = (
    "https://1drv.ms/x/c/e0a0290b0b3922a1/"
    "IQCYGmNIT3QJQLtGaZjj15TPAbOngjBaOcdtT6d11_9eL68?e=dAWT5T"
)

_REQUIRED_KEYS = ["MS_TENANT_ID", "MS_CLIENT_ID", "MS_CLIENT_SECRET"]


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return round(float(value), 2)
        except (ValueError, TypeError):
            return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace(" ", "").replace("Bs", "").replace("$", "")
    # Handle es-VE style "1.234,56"
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


class OneDriveService:
    def __init__(self):
        self.worksheet = os.environ.get("MS_WORKSHEET", WORKSHEET)
        self.table = os.environ.get("MS_TABLE", "_202606_Precios")
        self.share_url = os.environ.get("ONEDRIVE_SHARE_URL", DEFAULT_SHARE_URL)

    def is_configured(self) -> bool:
        return all(os.environ.get(k) for k in _REQUIRED_KEYS)

    def missing_setup(self):
        return [k for k in _REQUIRED_KEYS if not os.environ.get(k)]

    def _share_id(self) -> str:
        encoded = base64.urlsafe_b64encode(self.share_url.encode("utf-8")).decode("ascii")
        return "u!" + encoded.rstrip("=")

    async def _token(self, client: httpx.AsyncClient) -> str:
        tenant = os.environ["MS_TENANT_ID"]
        url = f"https://login.microsoftonline.com/{tenant}/oauth2/v2.0/token"
        data = {
            "grant_type": "client_credentials",
            "client_id": os.environ["MS_CLIENT_ID"],
            "client_secret": os.environ["MS_CLIENT_SECRET"],
            "scope": "https://graph.microsoft.com/.default",
        }
        r = await client.post(url, data=data)
        if r.status_code != 200:
            raise RuntimeError(f"token_failed: {r.status_code}: {r.text[:300]}")
        return r.json()["access_token"]

    async def _download(self, client: httpx.AsyncClient, token: str) -> bytes:
        sid = self._share_id()
        headers = {"Authorization": f"Bearer {token}"}
        item = await client.get(f"{GRAPH}/shares/{sid}/driveItem", headers=headers)
        if item.status_code != 200:
            raise RuntimeError(f"share_resolution_failed: {item.status_code}: {item.text[:300]}")
        content = await client.get(f"{GRAPH}/shares/{sid}/driveItem/content", headers=headers)
        if content.status_code != 200:
            raise RuntimeError(f"download_failed: {content.status_code}: {content.text[:300]}")
        return content.content

    def _map_row(self, row, id_idx, price_idx):
        def cell(i):
            return row[i] if i is not None and i < len(row) else None

        sku_raw = cell(id_idx["sku"])
        if sku_raw is None or str(sku_raw).strip() == "":
            return None
        rin_raw = cell(id_idx["rin"])
        try:
            rin = int(float(rin_raw)) if rin_raw not in (None, "") else 0
        except (ValueError, TypeError):
            rin = str(rin_raw).strip()
        prices = {key: _to_number(cell(i)) for key, i in price_idx}
        return {
            "sku": str(sku_raw).strip(),
            "rin": rin,
            "marca": str(cell(id_idx["marca"]) or "").strip(),
            "descripcion": str(cell(id_idx["descripcion"]) or "").strip(),
            "prices": prices,
        }

    def _parse(self, content: bytes):
        # Non read-only so table metadata is available (openpyxl skips tables in
        # read_only mode). Price lists are small, so this is fine.
        wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
        try:
            if self.worksheet not in wb.sheetnames:
                raise RuntimeError(f"worksheet_not_found:{self.worksheet}")
            ws = wb[self.worksheet]

            id_idx = {field: col_letter_to_index(letter) for letter, field in IDENTITY_COLUMNS.items()}
            price_idx = [(c["key"], col_letter_to_index(c["letter"])) for c in PRICE_COLUMNS]

            # Prefer the named Excel table over a raw range.
            data_start = 2  # default: assume header on row 1
            data_end = ws.max_row
            tables = getattr(ws, "tables", {}) or {}
            table_obj = None
            if hasattr(tables, "items"):
                for name, t in tables.items():
                    if str(name).replace(" ", "") == self.table.replace(" ", ""):
                        table_obj = t
                        break
                if table_obj is None and len(tables):
                    table_obj = next(iter(tables.values()))
            if table_obj is not None and getattr(table_obj, "ref", None):
                import re
                start, end = table_obj.ref.split(":")
                sr = int(re.match(r"[A-Z]+(\d+)", start).group(1))
                er = int(re.match(r"[A-Z]+(\d+)", end).group(1))
                data_start = sr + 1  # skip table header row
                data_end = er
                logger.info("Reading table %s range %s", self.table, table_obj.ref)

            products = []
            for row in ws.iter_rows(min_row=data_start, max_row=data_end, values_only=True):
                if row is None:
                    continue
                mapped = self._map_row(row, id_idx, price_idx)
                if mapped is not None:
                    products.append(mapped)
            return products
        finally:
            wb.close()

    async def fetch_rows(self):
        """Return mapped product rows from the live worksheet, or None if the
        Microsoft connection is not configured yet. Raises on connection error."""
        if not self.is_configured():
            logger.info("OneDrive not configured; mock fallback. Missing: %s", self.missing_setup())
            return None
        async with httpx.AsyncClient(timeout=90, follow_redirects=True) as client:
            token = await self._token(client)
            content = await self._download(client, token)
        return self._parse(content)

    async def sync(self, db, build_mock_products, normalize):
        """Full source synchronization (master only + scheduler)."""
        try:
            rows = await self.fetch_rows()
        except Exception as e:  # keep existing data on failure; report to caller
            logger.warning("OneDrive sync failed: %s", e)
            raise

        source = "onedrive:Precios Actual" if rows is not None else "mock"
        products = rows if rows is not None else build_mock_products()

        for p in products:
            p["search_blob"] = " ".join([
                normalize(p["sku"]), normalize(p["marca"]),
                normalize(p["descripcion"]), normalize(str(p["rin"])),
            ])

        if products:
            await db.products.delete_many({})
            await db.products.insert_many(products)

        return {"source": source, "row_count": len(products), "worksheet": self.worksheet}
