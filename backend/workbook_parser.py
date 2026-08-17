"""Shared workbook parser — maps an .xlsx (worksheet "Precios Actual", preferring
the named Excel table `_202606_Precios`) into product rows using the confirmed
column-letter mapping in schema.py. Used by every data provider so the mapping
and business logic stay identical regardless of source (Zoho / OneDrive / mock).
"""

import io
import re
import logging

from openpyxl import load_workbook

from schema import PRICE_COLUMNS, IDENTITY_COLUMNS, col_letter_to_index

logger = logging.getLogger("venege.parser")


def _to_number(value):
    if value is None:
        return None
    if isinstance(value, (int, float)):
        try:
            return round(float(value), 2)
        except (ValueError, TypeError):
            return None
    s = str(value).strip()
    if s == "":
        return None
    s = s.replace(" ", "").replace("Bs", "").replace("$", "")
    if "," in s and "." in s:
        s = s.replace(".", "").replace(",", ".")
    elif "," in s:
        s = s.replace(",", ".")
    try:
        return round(float(s), 2)
    except (ValueError, TypeError):
        return None


def _map_row(row, id_idx, price_idx):
    def cell(i):
        return row[i] if i is not None and i < len(row) else None

    sku_raw = cell(id_idx["sku"])
    if sku_raw is None or str(sku_raw).strip() == "":
        return None
    rin_raw = cell(id_idx["rin"])
    try:
        rin = int(float(rin_raw)) if rin_raw not in (None, "") else 0
    except (ValueError, TypeError):
        rin = str(rin_raw).strip()
    prices = {key: _to_number(cell(i)) for key, i in price_idx}
    return {
        "sku": str(sku_raw).strip(),
        "rin": rin,
        "marca": str(cell(id_idx["marca"]) or "").strip(),
        "descripcion": str(cell(id_idx["descripcion"]) or "").strip(),
        "prices": prices,
    }


def parse_workbook(content: bytes, worksheet: str, table: str, require_table: bool = True):
    """Parse xlsx bytes → list of product dicts. Prefers the named table.
    Raises if the required worksheet (or table, when require_table) is missing."""
    wb = load_workbook(io.BytesIO(content), read_only=False, data_only=True)
    try:
        if worksheet not in wb.sheetnames:
            raise RuntimeError(f"worksheet_not_found:{worksheet}")
        ws = wb[worksheet]

        id_idx = {field: col_letter_to_index(letter) for letter, field in IDENTITY_COLUMNS.items()}
        price_idx = [(c["key"], col_letter_to_index(c["letter"])) for c in PRICE_COLUMNS]

        # Resolve any header-matched columns (e.g. BF Goodrich at AM) by header
        # text for stability, falling back to the letter index.
        header = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None) or ()
        for c in PRICE_COLUMNS:
            match = c.get("header_match")
            if not match:
                continue
            for i, val in enumerate(header):
                if val is not None and match.upper() in str(val).upper():
                    price_idx = [(k, (i if k == c["key"] else idx)) for k, idx in price_idx]
                    break

        data_start = 2
        data_end = ws.max_row
        tables = getattr(ws, "tables", {}) or {}
        table_obj = None
        if hasattr(tables, "items"):
            for name, t in tables.items():
                if str(name).replace(" ", "") == table.replace(" ", ""):
                    table_obj = t
                    break
            if table_obj is None and not require_table and len(tables):
                table_obj = next(iter(tables.values()))
        if require_table and table_obj is None:
            raise RuntimeError(f"table_not_found:{table}")
        if table_obj is not None and getattr(table_obj, "ref", None):
            start, end = table_obj.ref.split(":")
            sr = int(re.match(r"[A-Z]+(\d+)", start).group(1))
            er = int(re.match(r"[A-Z]+(\d+)", end).group(1))
            data_start = sr + 1
            data_end = er
            logger.info("Reading table %s range %s", table, table_obj.ref)

        products = []
        for row in ws.iter_rows(min_row=data_start, max_row=data_end, values_only=True):
            if row is None:
                continue
            mapped = _map_row(row, id_idx, price_idx)
            if mapped is not None:
                products.append(mapped)
        return products
    finally:
        wb.close()
